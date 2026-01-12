
import pandas as pd
from pathlib import Path
from openpyxl.utils import get_column_letter

# Updated column structure - Main Header and Sub Header moved to end
COLUMNS = [
    'No.',
    'Published Date',
    'Closing Date',
    'Closing Time',
    'Date Detected',
    'ITQ/ITT',
    'Calling Entity',
    'Description',
    'Link',
    'Main Header',
    'Sub Header',
    'Sourcing Doc No.'
]

# Columns for Awards Worksheet
# Columns for Awards Worksheet
AWARDS_COLUMNS = [
    'No.',
    'Awarded Date',     # 1. Before Published Date
    'Published Date',   
    'Award Value',      # 2. After Awarded Date
    'Awarded To',       # 3. After Award Value
    'Closing Date',
    'Closing Time',
    'Date Detected',
    'ITQ/ITT',
    'Calling Entity',
    'Description',
    'Link',
    'Main Header',
    'Sub Header',
    'Sourcing Doc No.'
]

def adjust_column_widths(worksheet, dataframe):
    """
    Auto-adjust column widths to fit content
    """
    for idx, col in enumerate(dataframe.columns, 1):
        column_letter = get_column_letter(idx)
        
        # Get max length of content in this column
        max_length = 0
        
        # Check header length
        max_length = max(max_length, len(str(col)))
        
        # Check all cell values in this column
        for value in dataframe[col].astype(str):
            max_length = max(max_length, len(value))
        
        # Set width with extra padding to prevent premature wrapping
        # (max 100 to prevent extremely wide columns)
        adjusted_width = min(max_length + 5, 100)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def export_to_excel(items, path, metadata=None):
    """
    Export items to Excel with proper column structure and auto-width.
    Separates GeBIZ Awards into a different worksheet.
    Adds a Settings worksheet with run details.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    
    # 1. Split items into Opportunities and Awards
    opportunities = []
    awards = []
    
    for idx, item in enumerate(items, 1):
        item['No.'] = idx # Temporary re-indexing, will re-do per sheet
        if item.get('_is_award'):
            awards.append(item)
        else:
            opportunities.append(item)
            
    # Re-index
    for idx, item in enumerate(opportunities, 1): item['No.'] = idx
    for idx, item in enumerate(awards, 1): item['No.'] = idx
    
    # 2. Prepare DataFrames
    # Create copies to avoid mutating original items during mapping?
    # Better to map fields before DataFrame creation or rename columns?
    # Normalized items keys: 'published_date', 'closing_date', 'company', 'title' ... 
    # 'awarded_to', 'award_value', 'awarded_date_str'
    
    # Map Award-specific keys to Column Names on the fly for Awards
    processed_awards = []
    for a in awards:
        # Create a shallow copy to modify
        pa = a.copy()
        
        # Explicit Mapping from Normalized Keys
        pa['Awarded To'] = a.get('awarded_to') or "No Award"
        pa['Award Value'] = a.get('award_value', '')
        pa['Awarded Date'] = a.get('awarded_date_str', '') or a.get('Awarded Date', '')
        
        # Ensure Published Date is preserved (already in 'a' as 'Published Date')
        
        # 'awarded_date_str' needs to go into 'Awarded Date'
        # BUT prepare_dates expects 'Awarded Date' to possibly be datetime compatible
        # If 'awarded_date_str' is "05 Jan 2026", pandas to_datetime should handle it.
        
        # Ensure we don't have 'Published Date' conflicting if we want to drop it or leave it blank
        # The AWARDS_COLUMNS list doesn't include 'Published Date', so it will just be ignored by DataFrame constructor?
        # Yes, passing columns=... selects those keys.
        
        processed_awards.append(pa)

    # 2. Prepare DataFrames
    df_opps = pd.DataFrame(opportunities, columns=COLUMNS)
    df_awds = pd.DataFrame(processed_awards, columns=AWARDS_COLUMNS)
    
    # Format Date Columns logic (Convert to Datetime, Don't stringify yet)
    # Format Date Columns logic (Convert to Datetime, Don't stringify yet)
    # Include Awarded Date in date columns processing
    date_cols = ['Published Date', 'Closing Date', 'Date Detected', 'Awarded Date']
    
    def prepare_dates(df):
        for col in date_cols:
            if col in df.columns:
                # Convert to datetime objects, handling errors
                df[col] = pd.to_datetime(df[col], errors='coerce')
        return df

    df_opps = prepare_dates(df_opps)
    df_awds = prepare_dates(df_awds)
    
    # 3. Prepare Settings DataFrame
    settings_data = []
    if metadata:
        for k, v in metadata.items():
            settings_data.append({'Setting': k, 'Value': str(v)})
    
    df_settings = pd.DataFrame(settings_data)

    # Formatting Styles
    from openpyxl.styles import Border, Side, Alignment, Font
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    wrap_align = Alignment(wrap_text=True, vertical='top')
    
    def apply_standard_style(ws, df):
        # Apply style to all cells in data range
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = wrap_align
                
        # Header Style
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.border = thin_border
            
    def make_links_clickable(ws, df):
        if 'Link' not in df.columns: return
        link_col_idx = df.columns.get_loc('Link') + 1
        col_letter = get_column_letter(link_col_idx)
        for cell in ws[col_letter]:
            if cell.row > 1 and cell.value:
                # Naive check if it looks like a URL
                val = str(cell.value)
                if val.startswith('http'):
                    cell.hyperlink = val
                    cell.font = Font(color="0000FF", underline="single")
    
    print(f"Writing to {path}...")
    print(f"  Opportunities: {len(opportunities)}")
    print(f"  Awards: {len(awards)}")
    
    with pd.ExcelWriter(path, engine='openpyxl', datetime_format='dd/mm/yyyy') as xw:
        # Sheet 1: Opportunities
        df_opps.to_excel(xw, index=False, sheet_name='Opportunities')
        
        # Sheet 2: GeBIZ Awards
        if not df_awds.empty:
            df_awds.to_excel(xw, index=False, sheet_name='GeBIZ Awards')
            
        # Sheet 3: Settings
        if not df_settings.empty:
            df_settings.to_excel(xw, index=False, sheet_name='Settings')
        
        # Formatting - Opportunities
        ws_opps = xw.sheets['Opportunities']
        ws_opps.auto_filter.ref = ws_opps.dimensions
        
        # Apply Styles
        apply_standard_style(ws_opps, df_opps)
        make_links_clickable(ws_opps, df_opps)
        
        # Apply strict Date Formatting & Alignment
        left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        for col_idx, col_name in enumerate(df_opps.columns, 1):
             if col_name in date_cols:
                 col_letter = get_column_letter(col_idx)
                 for cell in ws_opps[col_letter]:
                     if cell.row > 1:
                         cell.number_format = 'dd/mm/yyyy'
                         cell.alignment = left_align
        
        # Auto-adjust column widths (Max 100 chars, including headers)
        adjust_column_widths(ws_opps, df_opps)
        
        # Formatting - Awards
        if not df_awds.empty:
             ws_awds = xw.sheets['GeBIZ Awards']
             ws_awds.auto_filter.ref = ws_awds.dimensions
             apply_standard_style(ws_awds, df_awds)
             make_links_clickable(ws_awds, df_awds)
             
             # Date Format
             for col_idx, col_name in enumerate(df_awds.columns, 1):
                 if col_name in date_cols:
                     col_letter = get_column_letter(col_idx)
                     for cell in ws_awds[col_letter]:
                         if cell.row > 1:
                             cell.number_format = 'dd/mm/yyyy'
             
             # Auto-width
             adjust_column_widths(ws_awds, df_awds)
             
             # Right Align 'Award Value' and 'Closing Time'
             right_align_cols = ['Award Value', 'Closing Time']
             right_align = Alignment(horizontal='right', vertical='top', wrap_text=True)
             
             for col_name in right_align_cols:
                 if col_name in df_awds.columns:
                     col_idx = df_awds.columns.get_loc(col_name) + 1
                     col_letter = get_column_letter(col_idx)
                     for cell in ws_awds[col_letter]:
                         if cell.row > 1:
                             cell.alignment = right_align
        
        # Formatting - Settings
        if not df_settings.empty:
             ws_set = xw.sheets['Settings']
             apply_standard_style(ws_set, df_settings)
             ws_set.column_dimensions['A'].width = 25
             ws_set.column_dimensions['B'].width = 80 # Wide for feeds list
    
    print(f"✓ Export complete.")

def append_to_tender_comb(items, tender_comb_path):
    """Append items to existing Tender Comb workbook"""
    if not Path(tender_comb_path).exists():
        print(f"Warning: Tender Comb file not found at {tender_comb_path}")
        return
    
    # Reindex items with row numbers
    for idx, item in enumerate(items, 1):
        item['No.'] = idx
    
    df = pd.DataFrame(items, columns=COLUMNS)
    
    # Convert dates to datetime objects
    date_cols = ['Published Date', 'Closing Date', 'Date Detected']
    for col in date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
    
    try:
        # Generate sheet name from date
        sheet_name = f"GeBIZ_{pd.to_datetime('today').strftime('%Y_%m_%d')}"
        if 'Date Detected' in df and not df.empty:
             try:
                 sheet_name = f"GeBIZ_{df['Date Detected'].iloc[0].strftime('%d_%m_%Y')}"
             except: pass

        with pd.ExcelWriter(tender_comb_path, engine='openpyxl', mode='a', if_sheet_exists='overlay', datetime_format='dd/mm/yyyy') as xw:
            df.to_excel(xw, index=False, sheet_name=sheet_name)
            
            # Get the worksheet
            worksheet = xw.sheets[sheet_name]
            
            # Apply Date Formatting
            for col_idx, col_name in enumerate(df.columns, 1):
                 if col_name in date_cols:
                     col_letter = get_column_letter(col_idx)
                     for cell in worksheet[col_letter]:
                         if cell.row > 1:
                             cell.number_format = 'dd/mm/yyyy'

            # Auto-adjust column widths
            adjust_column_widths(worksheet, df)
        
        print(f"✓ Appended {len(items)} items to {tender_comb_path} (sheet: {sheet_name})")
    except Exception as e:
        print(f"Error appending to Tender Comb: {e}")
